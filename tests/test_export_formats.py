"""Cross-format export tests for RIS, BibTeX, and XLSX.

These tests exercise the field mapping and keyword handling in
scopus_for_dobby.core.export with realistic DB-format article dicts.
No API calls are made.
"""

import pytest

from scopus_for_dobby.core import export as export_mod

# ── Fixtures ──────────────────────────────────────────────────────────────────

FULL_ARTICLE = {
    "title": "Deep Learning for Image Segmentation: A Survey",
    "first_author": "Kim J.",
    "all_authors": [
        {"name": "Kim J.", "auid": "12345678"},
        {"name": "Lee S.", "auid": "23456789"},
        {"name": "Park H.", "auid": "34567890"},
    ],
    "journal": "IEEE Transactions on Pattern Analysis",
    "volume": "46",
    "issue": "3",
    "pages": "1234-1256",
    "cover_date": "2024-03-15",
    "doi": "10.1109/TPAMI.2024.001234",
    "eid": "2-s2.0-85012345678",
    "scopus_id": "85012345678",
    "cited_by": 42,
    "open_access": True,
    "abstract": "A comprehensive survey of segmentation methods.",
    "keywords": "deep learning | segmentation | survey",
    "issn": "0162-8828",
    "source_type": "Journal",
    "affiliations": ["Seoul National University"],
}

# Sparse article: no volume, issue, pages, or doi.
SPARSE_ARTICLE = {
    "title": "Transformer Models in NLP",
    "first_author": "Smith A.",
    "all_authors": [{"name": "Smith A.", "auid": "98765432"}],
    "journal": "Nature Machine Intelligence",
    "volume": "",
    "issue": "",
    "pages": "",
    "cover_date": "2023-06-01",
    "doi": "",
    "eid": "2-s2.0-85098765432",
    "scopus_id": "85098765432",
    "cited_by": 128,
    "open_access": False,
    "abstract": "Transformers applied to language tasks.",
    "keywords": "",
    "issn": "",
    "source_type": "Journal",
    "affiliations": [],
}


def _parse_ris(text: str) -> list[tuple[str, str]]:
    """Parse RIS text into an ordered list of (tag, value) pairs."""
    pairs = []
    for line in text.splitlines():
        if "  - " in line:
            tag, _, value = line.partition("  - ")
            pairs.append((tag.strip(), value))
    return pairs


# ── RIS field mapping ─────────────────────────────────────────────────────────


class TestRISFieldMapping:
    def test_full_article_fields(self, tmp_path):
        out = str(tmp_path / "full.ris")
        export_mod.export_ris([FULL_ARTICLE], out)
        text = (tmp_path / "full.ris").read_text()
        pairs = _parse_ris(text)
        mapped = dict(pairs)

        # TY first, ER last.
        assert pairs[0] == ("TY", "JOUR")
        assert pairs[-1][0] == "ER"

        assert mapped["TI"] == "Deep Learning for Image Segmentation: A Survey"
        # Journal mapped to JO for JOUR type.
        assert mapped["JO"] == "IEEE Transactions on Pattern Analysis"
        assert mapped["PY"] == "2024"
        assert mapped["VL"] == "46"
        assert mapped["IS"] == "3"
        assert mapped["SP"] == "1234"
        assert mapped["EP"] == "1256"
        assert mapped["DO"] == "10.1109/TPAMI.2024.001234"
        assert mapped["AB"] == "A comprehensive survey of segmentation methods."
        assert mapped["SN"] == "0162-8828"

    def test_sparse_article_omits_missing_fields(self, tmp_path):
        out = str(tmp_path / "sparse.ris")
        export_mod.export_ris([SPARSE_ARTICLE], out)
        text = (tmp_path / "sparse.ris").read_text()
        tags = {tag for tag, _ in _parse_ris(text)}

        # Present fields.
        assert "TY" in tags
        assert "TI" in tags
        assert "JO" in tags
        assert "PY" in tags
        assert "ER" in tags

        # Missing volume/issue/pages/doi/issn -> no tags emitted.
        assert "VL" not in tags
        assert "IS" not in tags
        assert "SP" not in tags
        assert "EP" not in tags
        assert "DO" not in tags
        assert "SN" not in tags

    def test_multiple_authors_order_preserved(self, tmp_path):
        out = str(tmp_path / "authors.ris")
        export_mod.export_ris([FULL_ARTICLE], out)
        text = (tmp_path / "authors.ris").read_text()
        au_values = [v for tag, v in _parse_ris(text) if tag == "AU"]
        assert au_values == ["Kim J.", "Lee S.", "Park H."]

    def test_keywords_one_kw_line_each(self, tmp_path):
        out = str(tmp_path / "kw.ris")
        export_mod.export_ris([FULL_ARTICLE], out)
        text = (tmp_path / "kw.ris").read_text()
        kw_values = [v for tag, v in _parse_ris(text) if tag == "KW"]
        assert kw_values == ["deep learning", "segmentation", "survey"]

    def test_empty_keywords_no_kw_lines(self, tmp_path):
        out = str(tmp_path / "nokw.ris")
        export_mod.export_ris([SPARSE_ARTICLE], out)
        text = (tmp_path / "nokw.ris").read_text()
        kw_values = [v for tag, v in _parse_ris(text) if tag == "KW"]
        assert kw_values == []


# ── BibTeX entry validity ─────────────────────────────────────────────────────


class TestBibTeX:
    def test_entry_validity_basics(self, tmp_path):
        out = str(tmp_path / "out.bib")
        export_mod.export_bibtex([FULL_ARTICLE, SPARSE_ARTICLE], out)
        content = (tmp_path / "out.bib").read_text()

        # Balanced braces across the whole file.
        assert content.count("{") == content.count("}")

        # Two @article entries with non-empty cite keys.
        import re

        keys = re.findall(r"@\w+\{([^,]*),", content)
        assert len(keys) == 2
        for key in keys:
            assert key.strip() != ""

    def test_authors_joined_with_and(self, tmp_path):
        out = str(tmp_path / "auth.bib")
        export_mod.export_bibtex([FULL_ARTICLE], out)
        content = (tmp_path / "auth.bib").read_text()
        assert "Kim J. and Lee S. and Park H." in content

    def test_keywords_comma_joined(self, tmp_path):
        out = str(tmp_path / "kw.bib")
        export_mod.export_bibtex([FULL_ARTICLE], out)
        content = (tmp_path / "kw.bib").read_text()
        assert "deep learning, segmentation, survey" in content

    def test_empty_keywords_no_field(self, tmp_path):
        out = str(tmp_path / "nokw.bib")
        export_mod.export_bibtex([SPARSE_ARTICLE], out)
        content = (tmp_path / "nokw.bib").read_text()
        # No keywords field emitted, and no crash.
        assert "keywords" not in content


# ── XLSX content ──────────────────────────────────────────────────────────────


class TestXLSX:
    def _read_cells(self, path):
        """Return {header: value} from the single article data row."""
        import openpyxl

        wb = openpyxl.load_workbook(path)
        ws = wb.active
        headers = [c.value for c in ws[1]]
        values = [c.value for c in ws[2]]
        return dict(zip(headers, values, strict=False))

    def test_keywords_semicolon_joined(self, tmp_path):
        pytest.importorskip("openpyxl")
        out = str(tmp_path / "full.xlsx")
        export_mod.export_xlsx([FULL_ARTICLE], out)
        cells = self._read_cells(out)
        assert cells["Keywords"] == "deep learning; segmentation; survey"

    def test_empty_keywords_empty_cell(self, tmp_path):
        pytest.importorskip("openpyxl")
        # FULL_ARTICLE gives an abstract so the institutional tier (which
        # includes the Keywords column) is selected.
        article = dict(SPARSE_ARTICLE)
        article["abstract"] = "Has an abstract to trigger full-tier columns."
        out = str(tmp_path / "sparse.xlsx")
        export_mod.export_xlsx([article], out)
        cells = self._read_cells(out)
        # openpyxl reads an empty-string cell back as None.
        assert cells["Keywords"] in (None, "")


# ── Cross-format keyword consistency ──────────────────────────────────────────


class TestKeywordConsistency:
    def test_same_input_three_formats(self, tmp_path):
        article = dict(FULL_ARTICLE)
        article["keywords"] = "a | b | c"

        ris_out = str(tmp_path / "c.ris")
        bib_out = str(tmp_path / "c.bib")
        xlsx_out = str(tmp_path / "c.xlsx")

        export_mod.export_ris([article], ris_out)
        export_mod.export_bibtex([article], bib_out)

        # RIS -> three KW lines.
        kw_values = [v for tag, v in _parse_ris((tmp_path / "c.ris").read_text()) if tag == "KW"]
        assert kw_values == ["a", "b", "c"]

        # BibTeX -> "a, b, c".
        assert "a, b, c" in (tmp_path / "c.bib").read_text()

        # XLSX -> "a; b; c".
        pytest.importorskip("openpyxl")
        export_mod.export_xlsx([article], xlsx_out)
        import openpyxl

        wb = openpyxl.load_workbook(xlsx_out)
        ws = wb.active
        headers = [c.value for c in ws[1]]
        values = [c.value for c in ws[2]]
        cells = dict(zip(headers, values, strict=False))
        assert cells["Keywords"] == "a; b; c"
