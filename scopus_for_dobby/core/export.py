"""Export articles to XLSX and BibTeX formats.

Reuses field mapping logic from the original scopus_tools.py.
"""

import re
import unicodedata

# ── Field definitions ─────────────────────────────────────────────────────────

FREE_FIELDS = [
    ("eid",          "EID"),
    ("scopus_id",    "Scopus ID"),
    ("doi",          "DOI"),
    ("title",        "Title"),
    ("first_author", "First Author"),
    ("journal",      "Journal"),
    ("volume",       "Volume"),
    ("cover_date",   "Cover Date"),
    ("cited_by",     "Citations"),
    ("open_access",  "Open Access"),
]

INST_EXTRA = [
    ("all_authors_str", "All Authors"),
    ("abstract",        "Abstract"),
    ("keywords",        "Keywords"),
    ("issn",            "ISSN"),
    ("issue",           "Issue"),
    ("pages",           "Pages"),
    ("source_type",     "Type"),
]

FULL_FIELDS = FREE_FIELDS + INST_EXTRA


def _build_row(article: dict) -> dict:
    """Build a flat export row from a DB article."""
    # Format all authors as string
    all_authors = article.get("all_authors", [])
    if isinstance(all_authors, list) and all_authors:
        authors_str = "; ".join(
            a.get("name", "") if isinstance(a, dict) else str(a)
            for a in all_authors
        )
    else:
        authors_str = article.get("first_author", "")

    # Affiliations as string
    affs = article.get("affiliations", [])
    if isinstance(affs, list):
        if affs and isinstance(affs[0], dict):
            affs_str = "; ".join(a.get("name", "") for a in affs)
        else:
            affs_str = "; ".join(str(a) for a in affs)
    else:
        affs_str = ""

    oa = article.get("open_access", False)
    if isinstance(oa, bool):
        oa_str = "Yes" if oa else "No"
    else:
        oa_str = "Yes" if str(oa) in ("1", "True", "true") else "No"

    return {
        "eid": article.get("eid", ""),
        "scopus_id": article.get("scopus_id", ""),
        "doi": article.get("doi", ""),
        "title": article.get("title", ""),
        "first_author": article.get("first_author", ""),
        "all_authors_str": authors_str,
        "journal": article.get("journal", ""),
        "issn": article.get("issn", ""),
        "volume": article.get("volume", ""),
        "issue": article.get("issue", ""),
        "pages": article.get("pages", ""),
        "cover_date": article.get("cover_date", ""),
        "cited_by": article.get("cited_by", ""),
        "abstract": article.get("abstract", ""),
        "keywords": article.get("keywords", ""),
        "open_access": oa_str,
        "source_type": article.get("source_type", ""),
        "affiliations": affs_str,
        "_tags": ", ".join(article.get("_tags", [])),
        "_notes": article.get("_notes", ""),
    }


# ── XLSX export ───────────────────────────────────────────────────────────────

def export_xlsx(articles: list[dict], output_path: str) -> dict:
    """Export articles to an Excel file.

    Args:
        articles: List of article dicts (DB format).
        output_path: Output .xlsx file path.

    Returns:
        Export summary.
    """
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as err:
        raise RuntimeError("openpyxl not installed. Run: pip install openpyxl") from err

    # Auto-detect tier
    has_abstract = any(a.get("abstract") for a in articles)
    fields = FULL_FIELDS if has_abstract else FREE_FIELDS

    # Add tags and notes columns
    fields = fields + [("_tags", "Tags"), ("_notes", "Notes")]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Scopus Articles"

    # Header style
    hdr_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    hdr_fill = PatternFill(start_color="2E5EAA", end_color="2E5EAA", fill_type="solid")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    keys = [k for k, _ in fields]
    headers = [h for _, h in fields]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
    ws.row_dimensions[1].height = 22

    body_align = Alignment(vertical="top", wrap_text=True)
    for article in articles:
        row_data = _build_row(article)
        ws.append([str(row_data.get(k, "")) for k in keys])

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = body_align

    col_widths = {
        "EID": 22, "Scopus ID": 15, "DOI": 35,
        "Title": 52, "First Author": 22, "All Authors": 42,
        "Journal": 36, "ISSN": 12, "Volume": 8,
        "Issue": 8, "Pages": 12, "Cover Date": 12,
        "Citations": 10, "Open Access": 11,
        "Abstract": 60, "Keywords": 35, "Type": 16,
        "Tags": 20, "Notes": 30,
    }
    for col_idx, (_, hdr) in enumerate(fields, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(hdr, 20)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(output_path)
    tier = "institutional" if has_abstract else "free-tier"
    return {
        "exported": len(articles),
        "format": "xlsx",
        "tier": tier,
        "output": output_path,
    }


# ── BibTeX export ─────────────────────────────────────────────────────────────

BIBTEX_TYPE_MAP = {
    "Journal": "article",
    "Conference Proceeding": "inproceedings",
    "Book": "book",
    "Book Series": "incollection",
    "Trade Journal": "article",
}


def _ascii_slug(s: str) -> str:
    s = unicodedata.normalize("NFKD", str(s)).encode("ascii", "ignore").decode()
    return re.sub(r"[^A-Za-z0-9]", "", s)


def _make_bibtex_key(article: dict, used: dict) -> str:
    creator = article.get("first_author", "Unknown")
    if "," in creator:
        last = creator.split(",")[0].strip()
    else:
        parts = creator.split()
        last = parts[-1] if parts else "Unknown"
    last = _ascii_slug(last)
    year = str(article.get("cover_date", ""))[:4]
    title_word = _ascii_slug((article.get("title", "").split() or [""])[0])
    base = f"{last}{year}{title_word}"
    n = used.get(base, 0)
    key = base if n == 0 else f"{base}{chr(96 + n)}"
    used[base] = n + 1
    return key


def _escape_bib(s: str) -> str:
    return str(s).replace("{", r"\{").replace("}", r"\}").replace("&", r"\&")


def export_bibtex(articles: list[dict], output_path: str) -> dict:
    """Export articles to BibTeX format.

    Args:
        articles: List of article dicts (DB format).
        output_path: Output .bib file path.

    Returns:
        Export summary.
    """
    used_keys = {}
    bib_entries = []

    for article in articles:
        key = _make_bibtex_key(article, used_keys)
        source_type = article.get("source_type", "")
        etype = BIBTEX_TYPE_MAP.get(source_type, "misc")

        # Authors
        all_authors = article.get("all_authors", [])
        if isinstance(all_authors, list) and all_authors:
            authors_bib = " and ".join(
                a.get("name", "") if isinstance(a, dict) else str(a)
                for a in all_authors if (a.get("name") if isinstance(a, dict) else a)
            )
        else:
            authors_bib = article.get("first_author", "")

        src_field = "journal" if etype in ("article", "misc") else "booktitle"
        note = (
            f"Cited by {article.get('cited_by', '0')} (Scopus). "
            f"EID: {article.get('eid', '')}"
        )

        def field(name, value):
            v = str(value).strip() if value else ""
            return f"  {name:<12} = {{{_escape_bib(v)}}}," if v else ""

        keywords = article.get("keywords", "")
        if isinstance(keywords, str):
            keywords = keywords.replace(" | ", ", ")

        lines = [f"@{etype}{{{key},"]
        for f_str in [
            field("author", authors_bib),
            field("title", article.get("title", "")),
            field(src_field, article.get("journal", "")),
            field("year", str(article.get("cover_date", ""))[:4]),
            field("volume", article.get("volume", "")),
            field("number", article.get("issue", "")),
            field("pages", article.get("pages", "")),
            field("issn", article.get("issn", "")),
            field("doi", article.get("doi", "")),
            field("abstract", article.get("abstract", "")),
            field("keywords", keywords),
            field("note", note),
        ]:
            if f_str:
                lines.append(f_str)
        lines.append("}")
        bib_entries.append("\n".join(lines))

    with open(output_path, "w", encoding="utf-8") as f:
        f.write("\n\n".join(bib_entries) + "\n")

    return {
        "exported": len(articles),
        "format": "bibtex",
        "output": output_path,
    }
